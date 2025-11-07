#!/usr/bin/env python3
"""
IOC to Excel Template Processor - Notebook Version
Split into individual cells for Jupyter Notebook or sequential execution
"""

# =============================================================================
# CELL 1: Import Required Libraries and Check Dependencies
# =============================================================================
print("CELL 1: Importing libraries and checking dependencies...")

import os
import csv
import time
import io
import shutil
import traceback
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional, Any

# Check and import required packages
def check_package(module_name, package_name):
    try:
        module = __import__(module_name)
        print(f"  ✓ {package_name} is installed")
        return module
    except ImportError:
        print(f"  ✗ {package_name} is NOT installed")
        print(f"    Run: pip install {package_name}")
        return None

# Check required packages
print("\nChecking required packages:")
openpyxl = check_package('openpyxl', 'openpyxl')
msoffcrypto = check_package('msoffcrypto', 'msoffcrypto-tool')
requests = check_package('requests', 'requests')

if not all([openpyxl, msoffcrypto, requests]):
    print("\n⚠ WARNING: Some packages are missing. Install them before proceeding.")
else:
    print("\n✓ All required packages are installed")
    from openpyxl import load_workbook
    import msoffcrypto
    import requests

# =============================================================================
# CELL 2: Configuration Settings - UPDATE THESE VALUES
# =============================================================================
print("\nCELL 2: Configuration Settings")
print("=" * 50)

# REQUIRED: Update these paths for your environment
IOC_SOURCE_FOLDER = "C:/path/to/ioc/folder"  # Folder containing IOC CSV files
EXCEL_TEMPLATE_PATH = "C:/path/to/template.xlsx"  # Your password-protected template
OUTPUT_EXCEL_PATH = "C:/path/to/output/compiled_iocs.xlsx"  # Where to save results

# Excel Template Settings
EXCEL_PASSWORD = "your_password_here"  # Password for the template (leave empty if not protected)
EXCEL_SHEET_NAME = "Sheet1"  # Name of the sheet to add IOCs to
EXCEL_HEADER_ROW = 3  # Row number where headers are located

# VirusTotal API (Optional - for MD5 enrichment)
VT_API_KEY = ""  # Leave empty to skip VirusTotal enrichment
VT_API_URL = "https://www.virustotal.com/api/v3/files/"
VT_RATE_LIMIT = 15  # Seconds between requests (free tier: 4 req/min)

# Default IOC Values
DEFAULT_TLP = "TLP: AMBER+STRICT"
DEFAULT_CLASSIFICATION = "UNCLASSIFIED"

# Display current configuration
print("Current Configuration:")
print(f"  IOC Source: {IOC_SOURCE_FOLDER}")
print(f"  Template: {EXCEL_TEMPLATE_PATH}")
print(f"  Output: {OUTPUT_EXCEL_PATH}")
print(f"  Password Protected: {'Yes' if EXCEL_PASSWORD else 'No'}")
print(f"  Header Row: {EXCEL_HEADER_ROW}")
print(f"  VirusTotal: {'Enabled' if VT_API_KEY else 'Disabled'}")

if IOC_SOURCE_FOLDER == "C:/path/to/ioc/folder":
    print("\n⚠ WARNING: Using default paths. Update the configuration above!")

# =============================================================================
# CELL 3: Validate Configuration and Paths
# =============================================================================
print("\nCELL 3: Validating Configuration")
print("=" * 50)

validation_passed = True

# Check IOC source folder
print(f"\nChecking IOC Source Folder:")
print(f"  Path: {IOC_SOURCE_FOLDER}")
if os.path.exists(IOC_SOURCE_FOLDER):
    csv_count = len(list(Path(IOC_SOURCE_FOLDER).glob("*.csv")))
    print(f"  ✓ Folder exists ({csv_count} CSV files found)")
else:
    print(f"  ✗ Folder does not exist")
    validation_passed = False

# Check Excel template
print(f"\nChecking Excel Template:")
print(f"  Path: {EXCEL_TEMPLATE_PATH}")
if os.path.exists(EXCEL_TEMPLATE_PATH):
    file_size = os.path.getsize(EXCEL_TEMPLATE_PATH)
    print(f"  ✓ Template exists ({file_size:,} bytes)")
else:
    print(f"  ✗ Template file not found")
    validation_passed = False

# Check output directory
print(f"\nChecking Output Path:")
print(f"  Path: {OUTPUT_EXCEL_PATH}")
output_dir = os.path.dirname(OUTPUT_EXCEL_PATH)
if output_dir:
    if not os.path.exists(output_dir):
        try:
            os.makedirs(output_dir)
            print(f"  ✓ Created output directory: {output_dir}")
        except Exception as e:
            print(f"  ✗ Cannot create output directory: {e}")
            validation_passed = False
    else:
        print(f"  ✓ Output directory exists")

if not validation_passed:
    print("\n❌ Validation failed. Please fix the configuration and run again.")
else:
    print("\n✅ Configuration validated successfully")

# =============================================================================
# CELL 4: Define Excel Handler Functions
# =============================================================================
print("\nCELL 4: Defining Excel Handler Functions")
print("=" * 50)

def open_excel_workbook(file_path, password=None):
    """
    Open an Excel workbook, handling password protection if needed.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    
    print(f"  Opening: {file_path}")
    print(f"  File size: {os.path.getsize(file_path):,} bytes")
    
    # Handle password-protected files
    if password:
        try:
            print(f"  Attempting to decrypt password-protected file...")
            with open(file_path, 'rb') as f:
                office_file = msoffcrypto.OfficeFile(f)
                office_file.load_key(password=password)
                
                # Create decrypted copy in memory
                decrypted_stream = io.BytesIO()
                office_file.decrypt(decrypted_stream)
                
                # Load workbook from decrypted stream
                decrypted_stream.seek(0)
                workbook = load_workbook(decrypted_stream)
                print(f"  ✓ Successfully opened password-protected file")
                return workbook
                
        except Exception as e:
            print(f"  ✗ Failed to decrypt with password: {e}")
            print(f"  Attempting to open without password...")
    
    # Try opening without password
    try:
        workbook = load_workbook(file_path)
        print(f"  ✓ Successfully opened file (no password required)")
        print(f"  Available sheets: {', '.join(workbook.sheetnames)}")
        return workbook
    except Exception as e:
        raise Exception(f"Failed to open Excel file: {e}")

def add_iocs_to_worksheet(workbook, sheet_name, iocs, header_row=3):
    """
    Add IOCs to the specified worksheet.
    """
    try:
        # Get the target worksheet
        if sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            print(f"  Using sheet: '{sheet_name}'")
        else:
            print(f"  ⚠ Sheet '{sheet_name}' not found")
            print(f"  Available sheets: {', '.join(workbook.sheetnames)}")
            ws = workbook.active
            print(f"  Using active sheet: '{ws.title}'")
        
        # Read headers from the specified row
        headers = []
        header_map = {}  # Maps header text to column index
        
        print(f"  Reading headers from row {header_row}...")
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col_idx).value
            if cell_value:
                header_text = str(cell_value).strip()
                headers.append(header_text)
                header_map[header_text] = col_idx
        
        if not headers:
            print(f"  ✗ No headers found on row {header_row}")
            return False
        
        print(f"  Found {len(headers)} headers: {', '.join(headers)}")
        
        # Find the next empty row after headers
        next_row = header_row + 1
        
        # Check for existing data and find last row
        for row_idx in range(header_row + 1, ws.max_row + 1):
            for col_idx in range(1, min(ws.max_column + 1, 10)):
                if ws.cell(row=row_idx, column=col_idx).value:
                    next_row = row_idx + 1
                    break
        
        print(f"  Starting data insertion at row {next_row}")
        
        # Add each IOC to the worksheet
        rows_added = 0
        for ioc_data in iocs:
            row_has_data = False
            
            for header in headers:
                col_idx = header_map[header]
                
                # Try to match header with IOC data keys
                header_clean = header.rstrip(':')
                
                value = None
                # Try different key variations
                for key_variant in [header, header_clean, header + ':', header_clean + ':']:
                    if key_variant in ioc_data:
                        value = ioc_data[key_variant]
                        break
                
                if value:
                    ws.cell(row=next_row, column=col_idx, value=value)
                    row_has_data = True
            
            if row_has_data:
                rows_added += 1
                next_row += 1
        
        print(f"  ✓ Successfully added {rows_added} IOCs to Excel")
        return True
        
    except Exception as e:
        print(f"  ✗ Error adding IOCs: {e}")
        traceback.print_exc()
        return False

print("✓ Excel handler functions defined")

# =============================================================================
# CELL 5: Define IOC Processing Functions for Mandiant
# =============================================================================
print("\nCELL 5: Defining Mandiant IOC Processing Functions")
print("=" * 50)

def process_mandiant_csv(file_path):
    """
    Process Mandiant CSV file (25-*.csv format).
    """
    iocs = []
    
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.DictReader(f)
            
            print(f"  Processing: {os.path.basename(file_path)}")
            print(f"  Columns: {reader.fieldnames}")
            
            for row in reader:
                indicator_value = row.get('Indicator Value', '').strip()
                indicator_type = row.get('Indicator Type', '').strip()
                
                if not indicator_value or not indicator_type:
                    continue
                
                ioc_entry = {
                    'TLP:': DEFAULT_TLP,
                    'Classification:': DEFAULT_CLASSIFICATION,
                    'IOC': indicator_value,
                    'Association:': row.get('Association', ''),
                    'Type:': indicator_type.upper(),
                    'Note:': row.get('Note', ''),
                    'Source:': os.path.basename(file_path),
                    'Date:': datetime.now().strftime('%Y-%m-%d')
                }
                
                iocs.append(ioc_entry)
        
        print(f"  ✓ Extracted {len(iocs)} IOCs")
        
    except Exception as e:
        print(f"  ✗ Error processing {file_path}: {e}")
    
    return iocs

print("✓ Mandiant processing function defined")

# =============================================================================
# CELL 6: Define IOC Processing Functions for CrowdStrike
# =============================================================================
print("\nCELL 6: Defining CrowdStrike IOC Processing Functions")
print("=" * 50)

def process_crowdstrike_csv(file_path):
    """
    Process CrowdStrike CSV file (CSA-*.csv or CSIT-*.csv format).
    """
    iocs = []
    
    # CrowdStrike type mapping
    type_mapping = {
        'domain': 'DOMAIN',
        'ip_address': 'IPV4',
        'ipv4': 'IPV4',
        'ipv6': 'IPV6',
        'url': 'URL',
        'hash_sha1': 'SHA1',
        'hash_md5': 'MD5',
        'hash_sha256': 'SHA256',
        'email': 'EMAIL',
        'email_address': 'EMAIL'
    }
    
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.DictReader(f)
            
            print(f"  Processing: {os.path.basename(file_path)}")
            print(f"  Columns: {reader.fieldnames}")
            
            for row in reader:
                # Try different column name variations
                indicator_value = (row.get('indicator', '') or 
                                 row.get('Indicator', '') or 
                                 row.get('INDICATOR', '') or
                                 row.get('value', '')).strip()
                
                indicator_type = (row.get('type', '') or 
                                row.get('Type', '') or 
                                row.get('TYPE', '') or
                                row.get('indicator_type', '')).strip()
                
                if not indicator_value or not indicator_type:
                    continue
                
                # Map to standardized type
                standard_type = type_mapping.get(indicator_type.lower(), indicator_type.upper())
                
                ioc_entry = {
                    'TLP:': DEFAULT_TLP,
                    'Classification:': DEFAULT_CLASSIFICATION,
                    'IOC': indicator_value,
                    'Association:': row.get('association', ''),
                    'Type:': standard_type,
                    'Note:': row.get('note', '') or row.get('description', ''),
                    'Source:': os.path.basename(file_path),
                    'Date:': datetime.now().strftime('%Y-%m-%d')
                }
                
                iocs.append(ioc_entry)
        
        print(f"  ✓ Extracted {len(iocs)} IOCs")
        
    except Exception as e:
        print(f"  ✗ Error processing {file_path}: {e}")
    
    return iocs

print("✓ CrowdStrike processing function defined")

# =============================================================================
# CELL 7: Optional - Define VirusTotal Enrichment Function
# =============================================================================
print("\nCELL 7: Defining VirusTotal Enrichment Function (Optional)")
print("=" * 50)

def enrich_md5_with_virustotal(md5_hash, base_entry):
    """
    Query VirusTotal to get SHA1 and SHA256 for an MD5 hash.
    """
    if not VT_API_KEY:
        return []
    
    enriched = []
    headers = {"x-apikey": VT_API_KEY}
    
    try:
        response = requests.get(f"{VT_API_URL}{md5_hash}", headers=headers)
        
        if response.status_code == 200:
            data = response.json()
            attributes = data.get('data', {}).get('attributes', {})
            
            # Add SHA1 if available
            if sha1 := attributes.get('sha1'):
                sha1_entry = base_entry.copy()
                sha1_entry['IOC'] = sha1
                sha1_entry['Type:'] = 'SHA1'
                enriched.append(sha1_entry)
            
            # Add SHA256 if available
            if sha256 := attributes.get('sha256'):
                sha256_entry = base_entry.copy()
                sha256_entry['IOC'] = sha256
                sha256_entry['Type:'] = 'SHA256'
                enriched.append(sha256_entry)
            
            print(f"    ✓ Enriched MD5 {md5_hash[:8]}... with {len(enriched)} hashes")
        
        # Rate limiting for free tier
        time.sleep(VT_RATE_LIMIT)
        
    except Exception as e:
        print(f"    ✗ VirusTotal enrichment failed: {e}")
    
    return enriched

if VT_API_KEY:
    print("✓ VirusTotal enrichment function defined and enabled")
else:
    print("⚠ VirusTotal enrichment disabled (no API key)")

# =============================================================================
# CELL 8: Scan for IOC Files
# =============================================================================
print("\nCELL 8: Scanning for IOC Files")
print("=" * 50)

# Initialize file lists
mandiant_files = []
crowdstrike_files = []
other_files = []

if os.path.exists(IOC_SOURCE_FOLDER):
    folder = Path(IOC_SOURCE_FOLDER)
    
    # Find Mandiant files (25-*.csv)
    mandiant_files = list(folder.glob("25-*.csv"))
    
    # Find CrowdStrike files (CSA-*.csv or CSIT-*.csv)
    crowdstrike_files = (list(folder.glob("CSA-*.csv")) + 
                         list(folder.glob("CSIT-*.csv")))
    
    # Find other CSV files
    all_csv = list(folder.glob("*.csv"))
    other_files = [f for f in all_csv 
                   if f not in mandiant_files 
                   and f not in crowdstrike_files]
    
    # Report findings
    print(f"\nMandiant Files Found: {len(mandiant_files)}")
    for f in mandiant_files:
        print(f"  • {f.name}")
    
    print(f"\nCrowdStrike Files Found: {len(crowdstrike_files)}")
    for f in crowdstrike_files:
        print(f"  • {f.name}")
    
    if other_files:
        print(f"\nOther CSV Files (will be skipped): {len(other_files)}")
        for f in other_files[:5]:
            print(f"  • {f.name}")
        if len(other_files) > 5:
            print(f"  ... and {len(other_files) - 5} more")
    
    total_files = len(mandiant_files) + len(crowdstrike_files)
    print(f"\nTotal IOC files to process: {total_files}")
    
    if total_files == 0:
        print("\n⚠ No recognized IOC files found!")
        print("  Expected patterns: 25-*.csv, CSA-*.csv, CSIT-*.csv")
else:
    print(f"✗ IOC source folder does not exist: {IOC_SOURCE_FOLDER}")

# =============================================================================
# CELL 9: Extract IOCs from All Files
# =============================================================================
print("\nCELL 9: Extracting IOCs from Files")
print("=" * 50)

all_iocs = []

# Process Mandiant files
if mandiant_files:
    print("\nProcessing Mandiant files:")
    for csv_file in mandiant_files:
        iocs = process_mandiant_csv(str(csv_file))
        all_iocs.extend(iocs)
        
        # Add VirusTotal enrichment for MD5 hashes if enabled
        if VT_API_KEY:
            md5_iocs = [ioc for ioc in iocs if ioc.get('Type:', '').upper() == 'MD5']
            for md5_ioc in md5_iocs:
                enriched = enrich_md5_with_virustotal(md5_ioc['IOC'], md5_ioc)
                all_iocs.extend(enriched)

# Process CrowdStrike files
if crowdstrike_files:
    print("\nProcessing CrowdStrike files:")
    for csv_file in crowdstrike_files:
        iocs = process_crowdstrike_csv(str(csv_file))
        all_iocs.extend(iocs)

print(f"\n✓ Total IOCs extracted: {len(all_iocs)}")

# Show IOC type distribution
if all_iocs:
    type_counts = {}
    for ioc in all_iocs:
        ioc_type = ioc.get('Type:', 'Unknown')
        type_counts[ioc_type] = type_counts.get(ioc_type, 0) + 1
    
    print("\nIOC Types Distribution:")
    for ioc_type, count in sorted(type_counts.items()):
        print(f"  • {ioc_type}: {count}")
    
    # Show sample IOC
    print("\nSample IOC entry:")
    sample = all_iocs[0]
    for key, value in sample.items():
        print(f"  {key} {value}")
else:
    print("\n⚠ No IOCs were extracted")

# =============================================================================
# CELL 10: Prepare Output File
# =============================================================================
print("\nCELL 10: Preparing Output File")
print("=" * 50)

output_prepared = False

if all_iocs and os.path.exists(EXCEL_TEMPLATE_PATH):
    try:
        # Copy template to output location
        if os.path.abspath(EXCEL_TEMPLATE_PATH) != os.path.abspath(OUTPUT_EXCEL_PATH):
            shutil.copy2(EXCEL_TEMPLATE_PATH, OUTPUT_EXCEL_PATH)
            print(f"✓ Template copied to: {OUTPUT_EXCEL_PATH}")
            output_prepared = True
        else:
            # Create backup if paths are the same
            backup_path = OUTPUT_EXCEL_PATH + ".backup"
            shutil.copy2(EXCEL_TEMPLATE_PATH, backup_path)
            print(f"✓ Backup created: {backup_path}")
            output_prepared = True
            
    except Exception as e:
        print(f"✗ Failed to prepare output file: {e}")
else:
    if not all_iocs:
        print("⚠ No IOCs to process")
    if not os.path.exists(EXCEL_TEMPLATE_PATH):
        print(f"✗ Template file not found: {EXCEL_TEMPLATE_PATH}")

# =============================================================================
# CELL 11: Open Excel Template and Add IOCs
# =============================================================================
print("\nCELL 11: Processing Excel Template")
print("=" * 50)

if output_prepared and all_iocs:
    try:
        # Open the workbook
        print("Opening Excel template...")
        workbook = open_excel_workbook(
            OUTPUT_EXCEL_PATH,
            EXCEL_PASSWORD if EXCEL_PASSWORD != "your_password_here" else None
        )
        
        # Add IOCs to the sheet
        print(f"\nAdding {len(all_iocs)} IOCs to Excel...")
        success = add_iocs_to_worksheet(
            workbook,
            EXCEL_SHEET_NAME,
            all_iocs,
            EXCEL_HEADER_ROW
        )
        
        if success:
            # Save the workbook
            print("\nSaving Excel file...")
            workbook.save(OUTPUT_EXCEL_PATH)
            workbook.close()
            
            file_size = os.path.getsize(OUTPUT_EXCEL_PATH)
            print(f"✓ File saved successfully")
            print(f"  Size: {file_size:,} bytes")
            print(f"  Path: {OUTPUT_EXCEL_PATH}")
            
            # Note about password protection
            if EXCEL_PASSWORD and EXCEL_PASSWORD != "your_password_here":
                print("\n📝 IMPORTANT NOTE:")
                print("  The output file is NOT password-protected.")
                print("  To add password protection:")
                print("  1. Open the file in Excel")
                print("  2. File → Info → Protect Workbook → Encrypt with Password")
                print("  3. Enter your desired password")
        else:
            print("✗ Failed to add IOCs to Excel")
            workbook.close()
            
    except Exception as e:
        print(f"✗ Excel processing failed: {e}")
        traceback.print_exc()
else:
    print("⚠ Skipping Excel processing (no data or output not prepared)")

# =============================================================================
# CELL 12: Final Summary
# =============================================================================
print("\nCELL 12: Processing Summary")
print("=" * 50)

print("\nPROCESSING COMPLETE")
print("-" * 30)

if os.path.exists(OUTPUT_EXCEL_PATH) and all_iocs:
    print("✅ SUCCESS")
    print(f"  Files processed: {len(mandiant_files) + len(crowdstrike_files)}")
    print(f"  IOCs extracted: {len(all_iocs)}")
    print(f"  Output saved to: {OUTPUT_EXCEL_PATH}")
    
    # Type summary
    if all_iocs:
        type_counts = {}
        for ioc in all_iocs:
            ioc_type = ioc.get('Type:', 'Unknown')
            type_counts[ioc_type] = type_counts.get(ioc_type, 0) + 1
        
        print("\n  IOC Type Summary:")
        for ioc_type, count in sorted(type_counts.items()):
            print(f"    • {ioc_type}: {count}")
else:
    print("❌ PROCESSING INCOMPLETE")
    if not all_iocs:
        print("  No IOCs were extracted")
    if not os.path.exists(OUTPUT_EXCEL_PATH):
        print("  Output file was not created")

print("\nScript execution finished")
