# Cell 1: Import Required Libraries
import os
import csv
import requests
import time
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
import msoffcrypto
import io
import shutil

# Cell 2: VirusTotal API Configuration
VT_API_KEY = "YOUR_VIRUSTOTAL_API_KEY_HERE"
VT_API_URL = "https://www.virustotal.com/api/v3/files/"

# Cell 3: File Path Configuration
# Path to the folder containing IOC CSV files to process
IOC_SOURCE_FOLDER = "C:/path/to/ioc/folder"  # UPDATE THIS: Folder with your CSV files

# Path to your Excel template file
EXCEL_TEMPLATE_PATH = "C:/path/to/template/template.xlsx"  # UPDATE THIS: Your Excel template

# Path where the updated Excel file will be saved
OUTPUT_EXCEL_PATH = "C:/path/to/output/compiled_iocs.xlsx"  # UPDATE THIS: Where to save results

# Excel Configuration
EXCEL_PASSWORD = ""  # Leave empty if template is not password protected
EXCEL_SHEET_NAME = "Sheet1"  # Name of the sheet where IOCs should be added
EXCEL_HEADER_ROW = 3  # Row number where headers are located

# Cell 4: Function to Open Excel (with or without password)
def open_excel_file(file_path, password=None):
    """
    Open an Excel file, handling both password-protected and unprotected files.
    """
    try:
        # Check if file exists
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        # If password is provided and not empty, try to open as protected
        if password:
            try:
                print(f"  Attempting to open password-protected file...")
                with open(file_path, 'rb') as f:
                    office_file = msoffcrypto.OfficeFile(f)
                    office_file.load_key(password=password)
                    
                    # Create a decrypted copy in memory
                    decrypted = io.BytesIO()
                    office_file.decrypt(decrypted)
                    
                    # Load the workbook from the decrypted stream
                    decrypted.seek(0)
                    workbook = load_workbook(decrypted)
                    print(f"  Successfully opened password-protected file")
                    return workbook
            except Exception as e:
                print(f"  Could not open with password: {e}")
                print(f"  Attempting to open without password...")
        
        # Try opening without password
        workbook = load_workbook(file_path)
        print(f"  Successfully opened file (no password required)")
        return workbook
            
    except Exception as e:
        print(f"Error opening Excel file: {e}")
        raise

# Cell 5: Function to Query VirusTotal
def get_vt_hashes(md5_hash):
    """
    Query VirusTotal API to get SHA1 and SHA256 for a given MD5 hash.
    Returns a dict with sha1 and sha256, or None if not found.
    """
    # Skip if API key not configured
    if VT_API_KEY == "YOUR_VIRUSTOTAL_API_KEY_HERE":
        return None
        
    headers = {
        "x-apikey": VT_API_KEY
    }
    
    try:
        response = requests.get(f"{VT_API_URL}{md5_hash}", headers=headers)
        
        if response.status_code == 200:
            data = response.json()
            attributes = data.get('data', {}).get('attributes', {})
            return {
                'sha1': attributes.get('sha1', ''),
                'sha256': attributes.get('sha256', '')
            }
        elif response.status_code == 404:
            print(f"    MD5 {md5_hash} not found in VirusTotal")
            return None
        else:
            print(f"    VirusTotal API error for {md5_hash}: {response.status_code}")
            return None
            
    except Exception as e:
        print(f"    Error querying VirusTotal for {md5_hash}: {e}")
        return None
    
    # Rate limiting: VirusTotal free tier allows 4 requests per minute
    time.sleep(15)  # 15 seconds between requests = 4 per minute

# Cell 6: Function to Process Mandiant CSV Files
def process_mandiant_csv(file_path):
    """
    Process a Mandiant CSV file and extract IOCs.
    Returns a list of IOC dictionaries.
    """
    iocs = []
    
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.DictReader(f)
            
            # Debug: Print column headers
            print(f"  Column headers found: {reader.fieldnames}")
            
            row_count = 0
            for row in reader:
                row_count += 1
                indicator_value = row.get('Indicator Value', '').strip()
                indicator_type = row.get('Indicator Type', '').strip()
                
                if not indicator_value or not indicator_type:
                    continue
                
                # Base IOC entry
                ioc_entry = {
                    'TLP:': 'TLP: AMBER+STRICT',
                    'Classification:': 'UNCLASSIFIED',
                    'IOC': indicator_value,
                    'Association:': '',
                    'Type:': indicator_type,
                    'Note:': '',
                    'Source:': os.path.basename(file_path),
                    'Date:': datetime.now().strftime('%Y-%m-%d')
                }
                
                iocs.append(ioc_entry)
                
                # If it's an MD5, query VirusTotal for SHA1 and SHA256
                if indicator_type.lower() == 'md5' or 'md5' in indicator_type.lower():
                    print(f"    Processing MD5: {indicator_value}")
                    vt_data = get_vt_hashes(indicator_value)
                    
                    if vt_data:
                        # Add SHA1 entry
                        if vt_data.get('sha1'):
                            sha1_entry = ioc_entry.copy()
                            sha1_entry['IOC'] = vt_data['sha1']
                            sha1_entry['Type:'] = 'SHA1'
                            iocs.append(sha1_entry)
                        
                        # Add SHA256 entry
                        if vt_data.get('sha256'):
                            sha256_entry = ioc_entry.copy()
                            sha256_entry['IOC'] = vt_data['sha256']
                            sha256_entry['Type:'] = 'SHA256'
                            iocs.append(sha256_entry)
            
            print(f"  Processed {row_count} rows, extracted {len(iocs)} IOCs")
    
    except Exception as e:
        print(f"  ERROR processing {file_path}: {e}")
    
    return iocs

# Cell 7: Function to Process CrowdStrike CSV Files
def process_crowdstrike_csv(file_path):
    """
    Process a CrowdStrike CSV file and extract IOCs.
    Returns a list of IOC dictionaries.
    """
    iocs = []
    
    # Mapping of CrowdStrike types to standardized types
    type_mapping = {
        'domain': 'DOMAIN',
        'ip_address': 'IPV4',
        'url': 'URL',
        'hash_sha1': 'SHA1',
        'hash_md5': 'MD5',
        'hash_sha256': 'SHA256',
        'email': 'EMAIL'
    }
    
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.DictReader(f)
            
            # Debug: Print column headers
            print(f"  Column headers found: {reader.fieldnames}")
            
            row_count = 0
            for row in reader:
                row_count += 1
                
                # Try different case variations for column names
                indicator_value = (row.get('indicator', '') or row.get('Indicator', '') or 
                                 row.get('INDICATOR', '')).strip()
                indicator_type = (row.get('type', '') or row.get('Type', '') or 
                                row.get('TYPE', '')).strip()
                
                if not indicator_value or not indicator_type:
                    continue
                
                # Convert CrowdStrike type to standardized type
                standardized_type = type_mapping.get(indicator_type.lower(), indicator_type.upper())
                
                # Base IOC entry
                ioc_entry = {
                    'TLP:': 'TLP: AMBER+STRICT',
                    'Classification:': 'UNCLASSIFIED',
                    'IOC': indicator_value,
                    'Association:': '',
                    'Type:': standardized_type,
                    'Note:': '',
                    'Source:': os.path.basename(file_path),
                    'Date:': datetime.now().strftime('%Y-%m-%d')
                }
                
                iocs.append(ioc_entry)
            
            print(f"  Processed {row_count} rows, extracted {len(iocs)} IOCs")
    
    except Exception as e:
        print(f"  ERROR processing {file_path}: {e}")
    
    return iocs

# Cell 8: Function to Add IOCs to Excel
def add_iocs_to_excel(workbook, sheet_name, iocs, header_row=3):
    """
    Add IOCs to an existing Excel worksheet.
    Headers are expected to be on the specified row (default: row 3).
    
    Args:
        workbook: The Excel workbook object
        sheet_name: Name of the sheet to add IOCs to
        iocs: List of IOC dictionaries to add
        header_row: Row number where headers are located (default: 3)
    """
    try:
        # Get the worksheet
        if sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
        else:
            print(f"  Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}")
            print(f"  Using active sheet: {workbook.active.title}")
            ws = workbook.active
        
        # Get the headers from the specified row
        headers = []
        header_positions = {}  # Map header names to column positions
        
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value:
                headers.append(cell_value)
                header_positions[cell_value] = col
        
        if not headers:
            print(f"  ERROR: No headers found on row {header_row}")
            return False
        
        print(f"  Headers found on row {header_row}: {headers}")
        
        # Find the next empty row (starting from the row after headers)
        next_row = header_row + 1
        
        # Find the last row with data
        for row in range(header_row + 1, ws.max_row + 1):
            has_data = False
            for col in range(1, len(headers) + 1):
                if ws.cell(row=row, column=col).value:
                    has_data = True
                    last_data_row = row
            if has_data:
                next_row = row + 1
        
        print(f"  Starting to add IOCs at row {next_row}")
        
        # Add each IOC to the worksheet
        iocs_added = 0
        for ioc in iocs:
            for header in headers:
                col_idx = header_positions[header]
                
                # Try to match header with IOC dictionary key
                # Remove trailing colons from headers if present
                header_clean = header.rstrip(':')
                
                # Try different key variations
                value = None
                if header in ioc:
                    value = ioc[header]
                elif header_clean in ioc:
                    value = ioc[header_clean]
                elif header + ':' in ioc:
                    value = ioc[header + ':']
                elif header_clean + ':' in ioc:
                    value = ioc[header_clean + ':']
                
                if value:
                    ws.cell(row=next_row, column=col_idx, value=value)
            
            next_row += 1
            iocs_added += 1
        
        print(f"  Successfully added {iocs_added} IOCs to Excel")
        return True
        
    except Exception as e:
        print(f"  ERROR adding IOCs to Excel: {e}")
        import traceback
        traceback.print_exc()
        return False

# Cell 9: Main Function to Process All Files
def process_ioc_files_to_excel():
    """
    Main function that:
    1. Reads all CSV files from the IOC source folder
    2. Extracts IOCs from each file
    3. Opens the Excel template
    4. Adds all IOCs to the template
    5. Saves the result to the output path
    """
    
    print("="*70)
    print("IOC TO EXCEL COMPILER")
    print("="*70)
    
    # Validate paths
    print("\n1. VALIDATING PATHS:")
    print(f"   IOC Source Folder: {IOC_SOURCE_FOLDER}")
    print(f"   Excel Template: {EXCEL_TEMPLATE_PATH}")
    print(f"   Output File: {OUTPUT_EXCEL_PATH}")
    
    if not os.path.exists(IOC_SOURCE_FOLDER):
        print(f"\n❌ ERROR: IOC source folder not found: {IOC_SOURCE_FOLDER}")
        print("   Please update IOC_SOURCE_FOLDER in the script")
        return
    
    if not os.path.exists(EXCEL_TEMPLATE_PATH):
        print(f"\n❌ ERROR: Excel template not found: {EXCEL_TEMPLATE_PATH}")
        print("   Please update EXCEL_TEMPLATE_PATH in the script")
        return
    
    # Ensure output directory exists
    output_dir = os.path.dirname(OUTPUT_EXCEL_PATH)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"   Created output directory: {output_dir}")
    
    # Find all CSV files in the IOC folder
    print("\n2. SCANNING FOR IOC FILES:")
    folder = Path(IOC_SOURCE_FOLDER)
    
    # Find Mandiant files (25-XXXXXXXX.csv)
    mandiant_files = list(folder.glob("25-*.csv"))
    
    # Find CrowdStrike files (CSA-XXXXXX.csv or CSIT-XXXXX.csv)
    crowdstrike_files = list(folder.glob("CSA-*.csv")) + list(folder.glob("CSIT-*.csv"))
    
    # Find any other CSV files
    all_csv_files = list(folder.glob("*.csv"))
    other_files = [f for f in all_csv_files if f not in mandiant_files and f not in crowdstrike_files]
    
    print(f"   Found {len(mandiant_files)} Mandiant file(s)")
    print(f"   Found {len(crowdstrike_files)} CrowdStrike file(s)")
    print(f"   Found {len(other_files)} other CSV file(s)")
    
    total_files = len(mandiant_files) + len(crowdstrike_files)
    
    if total_files == 0:
        print(f"\n⚠ No recognized IOC files found in {IOC_SOURCE_FOLDER}")
        if other_files:
            print(f"   Found {len(other_files)} unrecognized CSV files - these will be skipped")
            print(f"   Files must match pattern: 25-*.csv, CSA-*.csv, or CSIT-*.csv")
        return
    
    # Process all CSV files and collect IOCs
    print("\n3. EXTRACTING IOCs FROM FILES:")
    all_iocs = []
    
    # Process Mandiant files
    for csv_file in mandiant_files:
        print(f"\n   Processing Mandiant file: {csv_file.name}")
        iocs = process_mandiant_csv(csv_file)
        all_iocs.extend(iocs)
        print(f"   → Extracted {len(iocs)} IOCs")
    
    # Process CrowdStrike files
    for csv_file in crowdstrike_files:
        print(f"\n   Processing CrowdStrike file: {csv_file.name}")
        iocs = process_crowdstrike_csv(csv_file)
        all_iocs.extend(iocs)
        print(f"   → Extracted {len(iocs)} IOCs")
    
    print(f"\n   TOTAL IOCs EXTRACTED: {len(all_iocs)}")
    
    if not all_iocs:
        print("\n⚠ No IOCs were extracted from the files")
        return
    
    # Copy template to output location first
    print("\n4. PREPARING EXCEL FILE:")
    try:
        print(f"   Copying template to output location...")
        shutil.copy2(EXCEL_TEMPLATE_PATH, OUTPUT_EXCEL_PATH)
        print(f"   Template copied successfully")
    except Exception as e:
        print(f"   ❌ ERROR copying template: {e}")
        return
    
    # Open the Excel file and add IOCs
    print("\n5. ADDING IOCs TO EXCEL:")
    try:
        print(f"   Opening Excel file...")
        workbook = open_excel_file(OUTPUT_EXCEL_PATH, EXCEL_PASSWORD)
        
        # Add IOCs to the worksheet
        print(f"   Adding {len(all_iocs)} IOCs to sheet '{EXCEL_SHEET_NAME}'...")
        success = add_iocs_to_excel(workbook, EXCEL_SHEET_NAME, all_iocs, EXCEL_HEADER_ROW)
        
        if success:
            # Save the workbook
            print(f"   Saving Excel file...")
            workbook.save(OUTPUT_EXCEL_PATH)
            workbook.close()
            
            print(f"\n✅ SUCCESS!")
            print(f"   → Added {len(all_iocs)} IOCs to Excel")
            print(f"   → Output saved to: {OUTPUT_EXCEL_PATH}")
            
            # If password was used, remind about re-protecting
            if EXCEL_PASSWORD:
                print(f"\n📝 NOTE: The output file is not password-protected.")
                print(f"   To add password protection:")
                print(f"   1. Open the file in Excel")
                print(f"   2. Go to File > Info > Protect Workbook > Encrypt with Password")
                print(f"   3. Enter your password")
        else:
            print("\n❌ Failed to add IOCs to Excel file")
            
    except Exception as e:
        print(f"\n❌ ERROR processing Excel file: {e}")
        import traceback
        traceback.print_exc()

# Cell 10: Main Execution
if __name__ == "__main__":
    # Check for required packages
    print("Checking required packages...")
    required_packages = {
        'openpyxl': 'openpyxl',
        'msoffcrypto': 'msoffcrypto-tool'
    }
    
    missing_packages = []
    for module, package in required_packages.items():
        try:
            __import__(module)
            print(f"  ✓ {package} is installed")
        except ImportError:
            print(f"  ✗ {package} is not installed")
            missing_packages.append(package)
    
    if missing_packages:
        print(f"\n⚠ Missing required packages: {', '.join(missing_packages)}")
        print(f"  Install them using: pip install {' '.join(missing_packages)}")
        response = input("\nDo you want to continue anyway? (y/n): ")
        if response.lower() != 'y':
            exit(1)
    
    # Check if paths are configured
    if IOC_SOURCE_FOLDER == "C:/path/to/ioc/folder":
        print("\n❌ ERROR: You need to configure the file paths in the script!")
        print("   Edit the following variables in Cell 3:")
        print("   - IOC_SOURCE_FOLDER: Path to folder containing your IOC CSV files")
        print("   - EXCEL_TEMPLATE_PATH: Path to your Excel template file")
        print("   - OUTPUT_EXCEL_PATH: Where you want the result saved")
        exit(1)
    
    # Check VirusTotal API
    if VT_API_KEY == "YOUR_VIRUSTOTAL_API_KEY_HERE":
        print("\n⚠ VirusTotal API key not configured")
        print("  MD5 hashes will NOT be enriched with SHA1/SHA256")
        print("  To enable: Get a free API key at https://www.virustotal.com/gui/join-us")
        print("  Then update VT_API_KEY in the script\n")
    
    # Run the main process
    process_ioc_files_to_excel()
    
    print("\n" + "="*70)
    print("SCRIPT EXECUTION COMPLETED")
    print("="*70)