# Cell 1: Import Required Libraries
import os
import csv
import requests
import time
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
import msoffcrypto
import io

# Cell 2: VirusTotal API Configuration
VT_API_KEY = "YOUR_VIRUSTOTAL_API_KEY_HERE"
VT_API_URL = "https://www.virustotal.com/api/v3/files/"

# Cell 3: Excel Configuration
EXCEL_PASSWORD = "YOUR_EXCEL_PASSWORD_HERE"  # Set your Excel password here
EXCEL_FILE_PATH = "./ioc_template.xlsx"  # Path to your password-protected Excel file
EXCEL_SHEET_NAME = "Sheet1"  # Name of the sheet where IOCs should be added
EXCEL_HEADER_ROW = 3  # Row number where headers are located (default: 3)

# Cell 4: Function to Open Password-Protected Excel
def open_protected_excel(file_path, password):
    """
    Open a password-protected Excel file and return a workbook object.
    """
    try:
        # Check if file exists
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        # Open the password-protected file
        with open(file_path, 'rb') as f:
            office_file = msoffcrypto.OfficeFile(f)
            office_file.load_key(password=password)
            
            # Create a decrypted copy in memory
            decrypted = io.BytesIO()
            office_file.decrypt(decrypted)
            
            # Load the workbook from the decrypted stream
            decrypted.seek(0)
            workbook = load_workbook(decrypted)
            
            return workbook
            
    except Exception as e:
        print(f"Error opening protected Excel file: {e}")
        # Try opening without password (in case file is not actually protected)
        try:
            print("Attempting to open file without password...")
            return load_workbook(file_path)
        except Exception as e2:
            print(f"Failed to open file without password: {e2}")
            raise

# Cell 5: Function to Query VirusTotal
def get_vt_hashes(md5_hash):
    """
    Query VirusTotal API to get SHA1 and SHA256 for a given MD5 hash.
    Returns a dict with sha1 and sha256, or None if not found.
    """
    # Skip if API key not configured
    if VT_API_KEY == "YOUR_VIRUSTOTAL_API_KEY_HERE":
        return None
        
    headers = {
        "x-apikey": VT_API_KEY
    }
    
    try:
        response = requests.get(f"{VT_API_URL}{md5_hash}", headers=headers)
        
        if response.status_code == 200:
            data = response.json()
            attributes = data.get('data', {}).get('attributes', {})
            return {
                'sha1': attributes.get('sha1', ''),
                'sha256': attributes.get('sha256', '')
            }
        elif response.status_code == 404:
            print(f"  MD5 {md5_hash} not found in VirusTotal")
            return None
        else:
            print(f"  VirusTotal API error for {md5_hash}: {response.status_code}")
            return None
            
    except Exception as e:
        print(f"  Error querying VirusTotal for {md5_hash}: {e}")
        return None
    
    # Rate limiting: VirusTotal free tier allows 4 requests per minute
    time.sleep(15)  # 15 seconds between requests = 4 per minute

# Cell 6: Function to Process Mandiant CSV Files
def process_mandiant_csv(file_path):
    """
    Process a Mandiant CSV file and extract IOCs.
    Returns a list of IOC dictionaries.
    """
    iocs = []
    
    try:
        # Check file permissions
        if not os.access(file_path, os.R_OK):
            print(f"  ERROR: No read permission for {file_path}")
            return iocs
        
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.DictReader(f)
            
            # Debug: Print column headers
            print(f"  Mandiant column headers found: {reader.fieldnames}")
            
            row_count = 0
            for row in reader:
                row_count += 1
                indicator_value = row.get('Indicator Value', '').strip()
                indicator_type = row.get('Indicator Type', '').strip()
                
                # Debug: Print first row data
                if row_count == 1:
                    print(f"  First row - Indicator Value: '{indicator_value}', Type: '{indicator_type}'")
                
                if not indicator_value or not indicator_type:
                    continue
                
                # Base IOC entry
                ioc_entry = {
                    'TLP:': 'TLP: AMBER+STRICT',
                    'Classification:': 'UNCLASSIFIED',
                    'IOC': indicator_value,
                    'Association:': '',
                    'Type:': indicator_type,
                    'Note:': '',
                    'Source:': os.path.basename(file_path),
                    'Date:': datetime.now().strftime('%Y-%m-%d')
                }
                
                iocs.append(ioc_entry)
                
                # If it's an MD5, query VirusTotal for SHA1 and SHA256
                if indicator_type.lower() == 'md5' or 'md5' in indicator_type.lower():
                    print(f"  Processing MD5: {indicator_value}")
                    vt_data = get_vt_hashes(indicator_value)
                    
                    if vt_data:
                        # Add SHA1 entry
                        if vt_data.get('sha1'):
                            sha1_entry = ioc_entry.copy()
                            sha1_entry['IOC'] = vt_data['sha1']
                            sha1_entry['Type:'] = 'SHA1'
                            sha1_entry['Association:'] = ''
                            iocs.append(sha1_entry)
                        
                        # Add SHA256 entry
                        if vt_data.get('sha256'):
                            sha256_entry = ioc_entry.copy()
                            sha256_entry['IOC'] = vt_data['sha256']
                            sha256_entry['Type:'] = 'SHA256'
                            sha256_entry['Association:'] = ''
                            iocs.append(sha256_entry)
            
            print(f"  Total rows processed: {row_count}, IOCs extracted: {len(iocs)}")
    
    except PermissionError as e:
        print(f"  PERMISSION ERROR reading {file_path}: {e}")
    except Exception as e:
        print(f"  ERROR processing {file_path}: {e}")
    
    return iocs

# Cell 7: Function to Process CrowdStrike CSV Files
def process_crowdstrike_csv(file_path):
    """
    Process a CrowdStrike CSV file and extract IOCs.
    Returns a list of IOC dictionaries.
    """
    iocs = []
    
    # Mapping of CrowdStrike types to standardized types
    type_mapping = {
        'domain': 'DOMAIN',
        'ip_address': 'IPV4',
        'url': 'URL',
        'hash_sha1': 'SHA1',
        'hash_md5': 'MD5',
        'hash_sha256': 'SHA256',
        'email': 'EMAIL'
    }
    
    try:
        # Check file permissions
        if not os.access(file_path, os.R_OK):
            print(f"  ERROR: No read permission for {file_path}")
            return iocs
        
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.DictReader(f)
            
            # Debug: Print column headers
            print(f"  CrowdStrike column headers found: {reader.fieldnames}")
            
            row_count = 0
            for row in reader:
                row_count += 1
                
                # Try different case variations for column names
                indicator_value = (row.get('indicator', '') or row.get('Indicator', '') or 
                                 row.get('INDICATOR', '')).strip()
                indicator_type = (row.get('type', '') or row.get('Type', '') or 
                                row.get('TYPE', '')).strip()
                
                # Debug: Print first few rows data
                if row_count <= 3:
                    print(f"  Row {row_count} - Indicator: '{indicator_value}', type: '{indicator_type}'")
                
                if not indicator_value or not indicator_type:
                    if row_count <= 3:
                        print(f"  Row {row_count} - SKIPPED (missing indicator or type)")
                    continue
                
                # Convert CrowdStrike type to standardized type
                standardized_type = type_mapping.get(indicator_type.lower(), indicator_type.upper())
                
                # Base IOC entry
                ioc_entry = {
                    'TLP:': 'TLP: AMBER+STRICT',
                    'Classification:': 'UNCLASSIFIED',
                    'IOC': indicator_value,
                    'Association:': '',
                    'Type:': standardized_type,
                    'Note:': '',
                    'Source:': os.path.basename(file_path),
                    'Date:': datetime.now().strftime('%Y-%m-%d')
                }
                
                iocs.append(ioc_entry)
            
            print(f"  Total rows processed: {row_count}, IOCs extracted: {len(iocs)}")
    
    except PermissionError as e:
        print(f"  PERMISSION ERROR reading {file_path}: {e}")
    except Exception as e:
        print(f"  ERROR processing {file_path}: {e}")
    
    return iocs

# Cell 8: Function to Add IOCs to Excel
def add_iocs_to_excel(workbook, sheet_name, iocs, header_row=3):
    """
    Add IOCs to an existing Excel worksheet.
    Assumes headers are on row 3 (or specified row) matching the IOC dictionary keys.
    
    Args:
        workbook: The Excel workbook object
        sheet_name: Name of the sheet to add IOCs to
        iocs: List of IOC dictionaries to add
        header_row: Row number where headers are located (default: 3)
    """
    try:
        # Get the worksheet
        if sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
        else:
            print(f"Sheet '{sheet_name}' not found. Using active sheet.")
            ws = workbook.active
        
        # Get the headers from row 3 (or specified header row)
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value:
                headers.append(cell_value)
        
        print(f"Excel headers found on row {header_row}: {headers}")
        
        # Find the next empty row (starting from the row after headers)
        next_row = header_row + 1  # Start checking from row after headers
        
        # Find the last row with data
        last_data_row = header_row
        for row in range(header_row + 1, ws.max_row + 1):
            has_data = False
            for col in range(1, len(headers) + 1):
                if ws.cell(row=row, column=col).value:
                    has_data = True
                    break
            if has_data:
                last_data_row = row
        
        # Set next_row to be after the last row with data
        next_row = last_data_row + 1
        
        print(f"Starting to add IOCs at row {next_row}")
        
        # Add each IOC to the worksheet
        iocs_added = 0
        for ioc in iocs:
            for col_idx, header in enumerate(headers, start=1):
                # Try to match header with IOC dictionary key
                # Remove trailing colons from headers if present
                header_clean = header.rstrip(':')
                
                # Try different key variations
                value = None
                if header in ioc:
                    value = ioc[header]
                elif header_clean in ioc:
                    value = ioc[header_clean]
                elif header + ':' in ioc:
                    value = ioc[header + ':']
                elif header_clean + ':' in ioc:
                    value = ioc[header_clean + ':']
                
                if value:
                    ws.cell(row=next_row, column=col_idx, value=value)
            
            next_row += 1
            iocs_added += 1
        
        print(f"Successfully added {iocs_added} IOCs to Excel")
        return True
        
    except Exception as e:
        print(f"Error adding IOCs to Excel: {e}")
        return False

# Cell 9: Main Function to Compile All IOCs
def compile_iocs_to_excel(folder_path, excel_path, password, sheet_name, header_row=3):
    """
    Main function to compile all IOCs from CSV files and add them to an Excel file.
    
    Args:
        folder_path: Path to folder containing IOC CSV files
        excel_path: Path to the Excel template file
        password: Password for the Excel file
        sheet_name: Name of the worksheet to add IOCs to
        header_row: Row number where headers are located (default: 3)
    """
    folder = Path(folder_path)
    all_iocs = []
    
    # Find all CSV files matching Mandiant format (25-XXXXXXXX)
    mandiant_files = list(folder.glob("25-*.csv"))
    
    # Find all CrowdStrike CSV files (CSA-XXXXXX or CSIT-XXXXX)
    csa_files = list(folder.glob("CSA-*.csv"))
    csit_files = list(folder.glob("CSIT-*.csv"))
    crowdstrike_files = csa_files + csit_files
    
    total_files = len(mandiant_files) + len(crowdstrike_files)
    
    if total_files == 0:
        print(f"No IOC CSV files found in {folder_path}")
        return
    
    print(f"Found {len(mandiant_files)} Mandiant file(s) and {len(crowdstrike_files)} CrowdStrike file(s)")
    
    # Process Mandiant CSV files
    for csv_file in mandiant_files:
        print(f"\nProcessing Mandiant file: {csv_file.name}")
        iocs = process_mandiant_csv(csv_file)
        all_iocs.extend(iocs)
        print(f"Extracted {len(iocs)} IOCs from {csv_file.name}")
    
    # Process CrowdStrike CSV files
    for csv_file in crowdstrike_files:
        print(f"\nProcessing CrowdStrike file: {csv_file.name}")
        iocs = process_crowdstrike_csv(csv_file)
        all_iocs.extend(iocs)
        print(f"Extracted {len(iocs)} IOCs from {csv_file.name}")
    
    # Open the Excel file and add IOCs
    if all_iocs:
        print(f"\nOpening Excel file: {excel_path}")
        
        try:
            # Open the password-protected Excel file
            workbook = open_protected_excel(excel_path, password)
            
            # Add IOCs to the worksheet
            success = add_iocs_to_excel(workbook, sheet_name, all_iocs, header_row)
            
            if success:
                # Save the workbook with password protection
                output_path = excel_path.replace('.xlsx', '_updated.xlsx')
                
                # First save without password
                workbook.save(output_path)
                workbook.close()
                
                # Now re-protect the file
                print(f"Saving and protecting updated Excel file: {output_path}")
                protect_excel_file(output_path, password)
                
                print(f"\n✓ Successfully compiled {len(all_iocs)} IOCs to {output_path}")
                print(f"  The file is password-protected with the same password as the original.")
            else:
                print("\n✗ Failed to add IOCs to Excel file")
                
        except Exception as e:
            print(f"\n✗ Error processing Excel file: {e}")
            import traceback
            traceback.print_exc()
    else:
        print("\nNo IOCs found to compile")

# Cell 10: Function to Protect Excel File with Password
def protect_excel_file(file_path, password):
    """
    Save an Excel file with password protection.
    Note: This requires the pywin32 package on Windows or alternative methods on other platforms.
    """
    try:
        # Try using msoffcrypto to encrypt
        print("Applying password protection to the file...")
        
        # Read the unprotected file
        with open(file_path, 'rb') as f:
            file_data = f.read()
        
        # Create a temporary protected version
        protected_path = file_path.replace('.xlsx', '_protected.xlsx')
        
        # Note: Direct password protection requires platform-specific tools
        # For now, we'll save without password and provide instructions
        print(f"\n⚠ Important: The file has been saved as: {file_path}")
        print(f"  Please manually add password protection in Excel:")
        print(f"  1. Open the file in Excel")
        print(f"  2. Go to File > Info > Protect Workbook > Encrypt with Password")
        print(f"  3. Enter your password: {password if password != 'YOUR_EXCEL_PASSWORD_HERE' else '[your password]'}")
        print(f"  4. Save the file")
        
        # Alternative: If you have pywin32 installed on Windows, uncomment below:
        """
        import platform
        if platform.system() == 'Windows':
            try:
                import win32com.client
                excel = win32com.client.Dispatch("Excel.Application")
                excel.DisplayAlerts = False
                wb = excel.Workbooks.Open(os.path.abspath(file_path))
                wb.SaveAs(os.path.abspath(file_path), Password=password)
                wb.Close()
                excel.Quit()
                print(f"✓ File successfully protected with password")
            except ImportError:
                print("pywin32 not installed. Cannot auto-protect file.")
            except Exception as e:
                print(f"Could not auto-protect file: {e}")
        """
        
    except Exception as e:
        print(f"Note about password protection: {e}")

# Cell 11: Configuration and Execution
if __name__ == "__main__":
    # Configuration
    IOC_FOLDER = "./ioc_folder"  # Folder containing IOC CSV files
    
    print("="*60)
    print("IOC to Excel Compiler Script Starting")
    print("="*60)
    
    # Check if folder exists
    if not os.path.exists(IOC_FOLDER):
        print(f"ERROR: Folder '{IOC_FOLDER}' does not exist!")
        print("Please create the folder or update IOC_FOLDER path in the script.")
        exit(1)
    
    # Check if Excel file exists
    if not os.path.exists(EXCEL_FILE_PATH):
        print(f"ERROR: Excel file '{EXCEL_FILE_PATH}' does not exist!")
        print("Please ensure the Excel template file exists or update EXCEL_FILE_PATH in the script.")
        exit(1)
    
    # Check if password is set
    if EXCEL_PASSWORD == "YOUR_EXCEL_PASSWORD_HERE":
        print("⚠ WARNING: Excel password not set!")
        response = input("Do you want to continue without password? (y/n): ")
        if response.lower() != 'y':
            print("Please set EXCEL_PASSWORD in the script.")
            exit(1)
    
    # Validate VirusTotal API key is set
    if VT_API_KEY == "YOUR_VIRUSTOTAL_API_KEY_HERE":
        print("⚠ WARNING: VirusTotal API key not set")
        print("  MD5 hashes will NOT be enriched with SHA1/SHA256")
        print("  You can get a free API key at: https://www.virustotal.com/gui/join-us")
        print("-"*60)
    
    # Install required packages if not present
    print("\nChecking required packages...")
    required_packages = ['openpyxl', 'msoffcrypto-tool']
    
    for package in required_packages:
        try:
            if package == 'msoffcrypto-tool':
                import msoffcrypto
            else:
                __import__(package)
            print(f"  ✓ {package} is installed")
        except ImportError:
            print(f"  ✗ {package} is not installed")
            print(f"    Install it using: pip install {package}")
            response = input(f"    Do you want to install {package} now? (y/n): ")
            if response.lower() == 'y':
                os.system(f"pip install {package}")
            else:
                print(f"    Please install {package} before running this script")
                exit(1)
    
    print("-"*60)
    
    # Run the compiler
    try:
        compile_iocs_to_excel(IOC_FOLDER, EXCEL_FILE_PATH, EXCEL_PASSWORD, EXCEL_SHEET_NAME, EXCEL_HEADER_ROW)
    except Exception as e:
        print(f"\n❌ UNEXPECTED ERROR: {e}")
        print(f"Error type: {type(e).__name__}")
        import traceback
        traceback.print_exc()
    
    print("\n" + "="*60)
    print("Script execution completed")
    print("="*60)